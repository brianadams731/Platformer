import json
from openpyxl import load_workbook

def convertToJSON(workbookname):
    workbook = load_workbook(filename = f'{workbookname}.xlsx', data_only=True)
    first_sheet = workbook.sheetnames[0]
    sheet = workbook[first_sheet]

    matrix = []

    for i in range(1, sheet.max_column+1):
        column = []
        for j in range(1, sheet.max_row+1):
            cell_obj = sheet.cell(row=j, column=i)
            if type(cell_obj.value) == int:
                column.append(cell_obj.value)
            else:
                column.append(0)
                print(f"{cell_obj.value} is not a number, will substitute with 0")
        matrix.append(column)


    data = {"mapData":matrix}

    with open(f'{workbookname}.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

while True:
    try:
        val = input("Workbook name: ")
        convertToJSON(val)
        print(f'{val}.json successfully created')
        break
    except:
        print("Failed to convert, check name")
